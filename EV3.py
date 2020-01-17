import re
import dns
from dns import resolver
import socket
import smtplib

from openpyxl import load_workbook

wb = load_workbook('Book1.xlsx', data_only=True)
sh = wb["Sheet1"]

for row in sh['A{}:A{}'.format(sh.min_row + 1, sh.max_row)]:
    for cell in row:
        try:
            wb.save('Book1.xlsx')
            addressToVerify = cell.value
            match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
            if match == None:
                print('Bad Syntax for ' + addressToVerify)

            resolver = dns.resolver.Resolver()
            records = dns.resolver.query('dnspython.org', 'MX')

            mxRecord = records[0].exchange
            mxRecord = str(mxRecord)

            # Get local server hostname
            host = socket.gethostname()

            # SMTP lib setup (use debug level for full output)
            server = smtplib.SMTP()
            server.set_debuglevel(0)

            # SMTP Conversation
            server.connect(mxRecord)
            server.helo(host)
            server.mail('surajk@lambdatest.com')
            print(addressToVerify)
            code, message = server.rcpt(str(addressToVerify))
            server.quit()
            print(code)
            # Assume 250 as Success
            # Assume 550 as Failure
            if code == 550:
                sh.cell(row=cell.row, column=2).value = "Soft"
            elif code == 250:
                sh.cell(row=cell.row, column=2).value = "Success"
            elif code == 520:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 521:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 522:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 531:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 545:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 553:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 421:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 450:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 451:
                sh.cell(row=cell.row, column=2).value="Soft1"
            elif code == 452:
                sh.cell(row=cell.row, column=2).value="Soft1"

            else:
                sh.cell(row=cell.row, column=2).value = "Fail"

        except Exception as e:
            print("error",e," for address ",addressToVerify)

wb.save('Book1.xlsx')
print("Done")
